from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from rich import print as rprint
from rich.table import Table
from rich.console import Console
import subprocess
import os

# Function to generate Fibonacci series for 30 dates
def generate_fibonacci_series():
    fib_series = [0, 1, 2]
    for i in range(3, 14):
        fib_series.append(fib_series[-1] + fib_series[-2])
    return fib_series

def days_from_june_first():
    start_date = datetime(2024, 6, 1)
    today_date = datetime.today()
    delta = today_date - start_date

    return delta.days

def next_empty_col(sheet, row_num)->int:
    col_num = 1
    while sheet.cell(row=row_num, column=col_num).value is not None:
        col_num += 1
    return col_num


# Function to write today's learning to Excel
def add_learning(topic):
    today = datetime.now().date()
    currDate = today
    fib_series = generate_fibonacci_series()

    if not os.path.exists('learning.xlsx'):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Revision'
        tracking_sheet = workbook.create_sheet('Track')
        tracking_sheet.append(['Date', 'Topic'])
        sheet.append(['Revision_Date', 'Topic', "Revision_number", 'Revision_status'])
    else:
        workbook = openpyxl.load_workbook('learning.xlsx')
        sheet = workbook['Revision']
        tracking_sheet = workbook['Track']
    
    row_num = days_from_june_first()+2

    for i, days in enumerate(fib_series):
        row_num+=(days+1)
        revision_date = currDate + timedelta(days=days+1)
        currDate = revision_date
        sheet.cell(row=row_num, column=1, value=revision_date)
        col_num = next_empty_col(sheet, row_num)
        sheet.cell(row=row_num, column=col_num, value=topic)
        sheet.cell(row=row_num, column=col_num+1, value=i+1)
        sheet.cell(row=row_num, column=col_num+2, value='Incomplete')
    
    tracking_sheet.append([today, topic])
    workbook.save('learning.xlsx')

def reload_todos(todos:list, sheet, row_today):
    todos.clear()
    for i, cell in enumerate(sheet[row_today][1::3]):
        if cell.value is not None:
            todos.append((cell.value, sheet[row_today][(3*i)+2].value, sheet[row_today][(3*i)+3].value))


def create_table():
    table = Table(title="Revision Today")
    table.add_column("SR.NO", justify="center", style="cyan", no_wrap=True)
    table.add_column("TOPIC", style="magenta")
    table.add_column("REVISION NO", justify="center", style="yellow")
    table.add_column("STATUS", justify="center", style="green")
    return table

# Function to show today's todo
def show_todos():
    if not os.path.exists('learning.xlsx'):
        print("No learning data found.")
        return

    workbook = openpyxl.load_workbook('learning.xlsx')
    sheet = workbook['Revision']
    row_today = days_from_june_first()+2
    todos = []
    reload_todos(todos, sheet, row_today)
    console = Console()
    t = len(todos)
    if todos:
        while True:
            table = create_table()
            print("Select Todo number to mark complete\nEnter 0 to exit\n")
            for i, todo in enumerate(todos):
                # print(f"{i+1} - Topic: {todo[0]}, Revision Number: {todo[1]}, Status: {todo[2]}\n")
                table.add_row(str(i+1),todo[0],str(todo[1]), todo[2])
            console.print(table)
            choice = int(input("Select Option: "))
            if choice==0:
                break
            if choice<=t:
                sheet.cell(row=row_today, column=(choice*3)+1, value='complete')
                reload_todos(todos, sheet, row_today)
            else:
                print("Invalid option, Please selection valid option")        
            
    else:
        print("No todos for today.")
    workbook.save('learning.xlsx')

def show_incomplete_todos():
    if not os.path.exists('learning.xlsx'):
        print("No learning data found.")
        return

    workbook = openpyxl.load_workbook('learning.xlsx')
    sheet = workbook['Revision']
    row_upto_yesterday = days_from_june_first()+1
    console = Console()
    Incomplete_todos = []
    for k, row in enumerate(sheet.iter_rows(min_row=2, max_row=row_upto_yesterday, values_only=True)):
        if row[1] is not None:
            for i in range(1, len(row), 3):
                if row[i] is not None and row[i+2]=='Incomplete':
                    Incomplete_todos.append((k+2, i+3, row[i], row[i+1], row[i+2]))
    t = len(Incomplete_todos)
    if Incomplete_todos:
        while True:
            table = create_table()
            print("Select Todo number to mark complete\nEnter 0 to exit\n")
            for i, todo in enumerate(Incomplete_todos):
                # print(f"{i+1} - Topic: {todo[2]}, Revision Number: {todo[3]}, Status: {todo[4]}\n")
                table.add_row(str(i+1),todo[2],str(todo[3]), todo[4])
            console.print(table)
            choice = int(input("Select Option: "))
            if choice==0:
                break
            if choice<t:
                sheet.cell(row=Incomplete_todos[choice-1][0], column=Incomplete_todos[choice-1][1], value='complete')
                Incomplete_todos.pop(choice-1)
            else:
                print("Invalid option, Please selection valid option")
    else:
        print("Horray!, You dont' have any incomplete todos remaining")
    workbook.save('learning.xlsx')


def run_command(command):
    result = subprocess.run(command, capture_output=True, text=True, shell=True)
    if result.returncode != 0:
        print(f"Error: {result.stderr}")
    else:
        print(result.stdout)

def sync_repo(remote_name='origin'):
    today = datetime.today()
    run_command(f'git pull {remote_name} main')
    run_command('git add .')
    run_command(f'git commit -m "sync {today}"')
    run_command(f'git push {remote_name} main')    

# Main function to interact with the user
def main():
    while True:
        rprint("\nMenu:")
        rprint("1. Add today's learning")
        rprint("2. Show today's Revisions")
        rprint("3. Show Past incomplete Revisions")
        rprint("4. Exit")
        rprint("5. Sync")
        choice = input("Enter your choice: ")

        if choice == '1':
            topics_input = input("Enter today's learning topics (comma-separated): ")
            topics = [topic.strip() for topic in topics_input.split(',')]
            for topic in topics:
                add_learning(topic) 
            print("Your today's learnings has been added")
        elif choice == '2':
            show_todos()
        elif choice=='3':
            show_incomplete_todos()
        elif choice == '4':
            break
        elif choice =='5':
            sync_repo()
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
